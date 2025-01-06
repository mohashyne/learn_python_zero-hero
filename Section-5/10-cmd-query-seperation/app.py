import openpyxl

my_xldoc = "Section-5/10-cmd-query-seperation/my_readings.xlsx"


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(my_xldoc)
sheet = wb['31-Jan-2022'] 


# command query separation  is a design principle that states that a method should either change the state of an object, or return some information about the object, but not both. 
# This principle is used to make the code easier to understand and maintain.
# Example of command query separation:

# command
# wb.create_sheet()  # create a new sheet  
# sheet.insert_rows(0, 2) # insert 2 rows at the beginning

# query
# print(sheet.max_row) # 38
# sheet.cell()

# violation of command query separation
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)  # this is a violation of command query separation, because we are changing the state of the object and also returning some information about the object
sheet.append([1, 2, 3, 4])           



wb.save("cmd _qur_reading.xlsx")

# The problem withe the above code is that we are changing the state of the object and also returning some information about the object in the same method. This makes the code harder to understand and maintain. by iterating over the rows and printing the values of the cells, and we only only have 5 rows in the original sheet, and the range is 1 to 10, so it will add 5 empty rows to the sheet. This is not what we want. We want to only print the values of the cells in the sheet. So we need to separate the command and query in the code.

# i.e 
numbers = [ 1, 2, 3, 4, 5]
#  if we  try to access the 6th element in the list, we will get an error
print(numbers[6])  # this will raise an error, because the index is out of range

# you can see that as we try to access the 6th element in the list, we get an error, and it dose not magically add the 6th element to the list. This is the same thing we want to do with the code above. We want to only print the values of the cells in the sheet, and not add empty rows to the sheet. So we need to separate the command and query in the code.

# so anytime we use the square bracket to access the value of a cell, a magic method called __getitem__ is called, and this method is responsible for returning the value of the cell. So we can override this method to separate the command and query in the code.