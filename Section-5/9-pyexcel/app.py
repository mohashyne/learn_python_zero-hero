"""Working with excel"""

import openpyxl

my_xldoc = "Section-5/9-pyexcel/my_readings.xlsx"


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(my_xldoc)
print("Sheet names in the workbook:", wb.sheetnames)  # Sheet names in the workbook: ['31-Jan-2022']


sheet = wb['31-Jan-2022']   

# # # our wb object has couple of methods we can use
# # wb.create_sheet("Sheet2") 
# # # OR use an index number to specify the position of the worksheet
# # # wb.insert_sheet(0, "Sheet1")
# wb.create_sheet("Sheet3", 0) 

cell = sheet["a1"]  
# print(cell.value) # FeederID
# print(cell.row) # 1
# print(cell.column) # 1
# print(cell.coordinate)  # A1
# print(sheet.max_row) # 36
# print(sheet.max_column) # 36

# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value) # this will print all the values

# alternatively we can use square bracket to access values of a cell
column = sheet["b"]
# print(column)

cells = sheet["a:z"]
print(cells)
        
        

# optionally we can use ,the first approach is shorter and easier to use,
# but this is useful when we are iterating over all the rows and columns
# and we want to dynamically various cells
# for row in sheet.iter_rows(values_only=True):
# celll = sheet.cell(row=1, column=1)
celll = sheet.cell(row=1, column=1)
print(celll.value) # FeederID


# we can also change the  value
# cell.value = "New Value"  # it could be string or number




# sheet has different methods like: 
# sheet.insert_rows(0, 2) # insert 2 rows at the beginning
# sheet.insert_cols(0, 2) # insert 2 columns at the beginning
# sheet.delete_rows(0, 2) # delete 2 rows at the beginning
# sheet.delete_cols(0, 2) # delete 2 columns at the beginning
# sheet.insert_rows(0, 2) # insert 2 rows at the beginning
# append
# sheet.append(["value1", "value2", "value3"]) # append a row