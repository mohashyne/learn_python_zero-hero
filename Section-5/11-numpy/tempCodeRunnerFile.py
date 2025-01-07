import openpyxl

my_xldoc = "Section-5/10-cmd-query-seperation/my_readings.xlsx"


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(my_xldoc)
sheet = wb['31-Jan-2022'] 


# command query separation  is a design principle that states that a method should either change the state of an object, or return some information about the object, but not both. 
# This principle is used to make the code easier to understand and maintain.
# Example of command query separation:

# command
# wb.create_sheet()  # create a new sheet  
# sheet.insert_rows(0, 2) # insert 2 rows at the beginning

# query
# print(sheet.max_row) # 38
# sheet.cell()

# violation of command query separation
for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)  # this is a violation of command query separation, because we are changing the state of the object and also returning some information about the object
sheet.append([1, 2, 3, 4])           



wb.save("cmd _qur_reading.xlsx")